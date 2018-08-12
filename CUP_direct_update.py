"""
Spyder Editor

This is a temporary script file.

银联地区码表更新 此处主要为excel处理包openpyxl

@author: 陈开锋
"""
from openpyxl import load_workbook# -*- coding: utf-8 -*-
from openpyxl import Workbook
wb = load_workbook('C:/Users/陈开锋/Desktop/py/1.xlsx')
a_sheet = wb.get_sheet_by_name('Sheet3')
ws = wb.active
wb2 = Workbook()
ws2 = wb2.active
ws2.title = '地区码表'
a_sheet2 = wb2.get_sheet_by_name('地区码表')
sheng=0
qu=0
shi=0
rw2=1
for rw in range(1,ws.max_row+1):#ws.max_row:
    for j in range(2,3):
        cells=ws.cell(row=rw,column=j)
        if cells.font.b == True:#省级代码
            sheng=sheng+1
            provincial=ws.cell(row=rw,column=3)
            provincial_code=provincial.value
            county_code=provincial_code
            municipal_code=provincial_code
        elif cells.alignment.horizontal!='center' and cells.alignment.indent==0.0 :#区级代码
            qu=qu+1
            municipal=ws.cell(row=rw,column=3)
            municipal_code=municipal.value
            county_code=municipal_code
        elif (cells.alignment.indent==2.0 and cells.alignment.horizontal=='left'):#市（县）级代码
            shi=shi+1
            county=ws.cell(row=rw,column=3)
            county_code=county.value
        column3=ws.cell(row=rw,column=3)
        if column3.value != None:
            #print(county_code,municipal_code,provincial_code,cells.value)
            sheet2=ws2.cell(row=rw2,column=2)
            sheet2.value=county_code
            sheet2=ws2.cell(row=rw2,column=3)
            sheet2.value=municipal_code
            sheet2=ws2.cell(row=rw2,column=4)
            sheet2.value=provincial_code
            sheet2=ws2.cell(row=rw2,column=5)
            sheet2.value=cells.value
            rw2=rw2+1 
wb2.save('C:/Users/陈开锋/Desktop/py/example.xlsx')
print("省：",sheng,"  地级市：",qu,"  县级市：",shi)